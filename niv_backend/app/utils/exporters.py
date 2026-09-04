"""Exportadores de datos NIV a formatos Excel (.xlsx) y DXF (grabado láser)."""
from __future__ import annotations

import io
from collections.abc import Sequence
from typing import TYPE_CHECKING

import ezdxf
import pandas as pd
from ezdxf.enums import TextEntityAlignment
from openpyxl.styles import Alignment, Font, PatternFill

if TYPE_CHECKING:
    from app.models import NIV

# Mapa de columna interna -> encabezado en español para el reporte Excel.
_EXCEL_COLUMNS: dict[str, str] = {
    "niv": "NIV",
    "fecha_generacion": "Fecha Generación",
    "tipo_remolque": "Tipo",
    "num_ejes": "Ejes",
    "capacidad": "Capacidad",
    "tipo_frenos": "Frenos",
    "version": "Versión",
    "anio_modelo": "Año",
    "linea_produccion": "Línea",
    "numero_serie": "Serie",
}

_CHAR_WIDTH_FACTOR = 0.6  # relación ancho/alto aproximada para fuente Arial
_ROW_HEIGHT_FACTOR = 2.0  # separación vertical entre NIV en exportación batch


def export_niv_to_excel(registros: Sequence["NIV"]) -> io.BytesIO:
    """Genera un archivo .xlsx con formato profesional a partir de una lista de registros NIV."""
    filas = [
        {
            "niv": r.niv,
            "fecha_generacion": r.fecha_generacion.strftime("%Y-%m-%d %H:%M:%S") if r.fecha_generacion else "",
            "tipo_remolque": r.tipo_remolque,
            "num_ejes": r.num_ejes,
            "capacidad": r.capacidad,
            "tipo_frenos": r.tipo_frenos,
            "version": r.version,
            "anio_modelo": r.anio_modelo,
            "linea_produccion": r.linea_produccion,
            "numero_serie": r.numero_serie,
        }
        for r in registros
    ]
    df = pd.DataFrame(filas, columns=list(_EXCEL_COLUMNS.keys())).rename(columns=_EXCEL_COLUMNS)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="NIV")
        worksheet = writer.sheets["NIV"]
        _style_header(worksheet, len(_EXCEL_COLUMNS))
        _autosize_columns(worksheet, df)
    buffer.seek(0)
    return buffer


def _style_header(worksheet, num_columns: int) -> None:
    """Aplica relleno azul y fuente blanca en negrita al encabezado del reporte."""
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_index in range(1, num_columns + 1):
        cell = worksheet.cell(row=1, column=col_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(worksheet, df: pd.DataFrame) -> None:
    """Ajusta el ancho de cada columna al contenido más largo."""
    for col_index, column in enumerate(df.columns, start=1):
        larguras = [len(str(column))] + [len(str(v)) for v in df[column]]
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_index).column_letter].width = max(larguras) + 4


def export_niv_to_dxf(
    nivs: Sequence[str],
    altura_texto: float,
    espaciado: float,
    fuente: str = "Arial",
    batch: bool = True,
) -> io.BytesIO:
    """Genera un archivo DXF (AutoCAD R2018) con los NIV para grabado láser.

    Cada carácter se dibuja como una entidad TEXT independiente, centrada en el
    origen (0,0), con capa "NIV" y color ACI 7 (negro). Si `batch` es False, solo
    se incluye el primer NIV de la lista.
    """
    nivs_a_exportar = list(nivs) if batch else list(nivs[:1])

    doc = ezdxf.new(dxfversion="R2018")

    if "NIV" not in doc.layers:
        doc.layers.new("NIV", dxfattribs={"color": 7})
    if fuente not in doc.styles:
        doc.styles.new(fuente, dxfattribs={"font": "arial.ttf"})

    msp = doc.modelspace()

    char_width = altura_texto * _CHAR_WIDTH_FACTOR
    row_height = altura_texto * _ROW_HEIGHT_FACTOR
    total_rows = len(nivs_a_exportar)
    start_y = ((total_rows - 1) * row_height) / 2

    for row_index, niv in enumerate(nivs_a_exportar):
        y = start_y - row_index * row_height
        total_width = len(niv) * char_width + max(len(niv) - 1, 0) * espaciado
        x = -total_width / 2
        for char in niv:
            center_x = x + char_width / 2
            text = msp.add_text(
                char,
                dxfattribs={
                    "layer": "NIV",
                    "height": altura_texto,
                    "color": 7,
                    "style": fuente,
                },
            )
            text.set_placement((center_x, y), align=TextEntityAlignment.MIDDLE_CENTER)
            x += char_width + espaciado

    text_stream = io.StringIO()
    doc.write(text_stream)
    return io.BytesIO(text_stream.getvalue().encode("utf-8"))
